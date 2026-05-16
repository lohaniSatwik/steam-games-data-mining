import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

wb = openpyxl.Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────
def hfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def bold_font(size=11, color='000000'):
    return Font(bold=True, size=size, color=color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row, headers, header_fill, font_color='FFFFFF'):
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill      = hfill(header_fill)
        cell.font      = Font(bold=True, size=11, color=font_color)
        cell.alignment = center()
        cell.border    = thin_border()

def write_row(ws, row, values, fill_hex=None, bold=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = center()
        cell.border    = thin_border()
        if fill_hex:
            cell.fill = hfill(fill_hex)
        if bold:
            cell.font = Font(bold=True, size=11)

def add_title(ws, row, text, ncols, fill_hex):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = hfill(fill_hex)
    cell.font      = Font(bold=True, size=13, color='FFFFFF')
    cell.alignment = center()
    cell.border    = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_finding_box(ws, row, text, ncols, height=60):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = hfill('EBF3FB')
    cell.font      = Font(italic=True, size=10, color='1F3864')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border    = thin_border()
    ws.row_dimensions[row].height = height


# =============================================================================
# SHEET 1 — Full Model Comparison
# =============================================================================
ws1 = wb.active
ws1.title = 'Model Comparison'
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 42

NCOLS1 = 10
add_title(ws1, 1,
    'IE500 Data Mining  |  Team 9 – Brewed Clusters  |  Full Model Comparison (0.5-Threshold Dataset)',
    NCOLS1, '1F3864')

HEADERS1 = [
    'Model',
    'CV Macro F1\n(mean)',
    'CV Std\n(+/-)',
    'Test\nMacro F1',
    'Good F1\n(test)',
    'Mixed F1\n(test)',
    'Bad F1\n(test)',
    'vs LR\nBaseline',
    'Rank',
    'Notes / Best Hyperparameters',
]
write_header(ws1, 2, HEADERS1, '2E75B6')

# (model, cv_mean, cv_std, test_f1, good_f1, mixed_f1, bad_f1, vs_lr, rank, notes)
models_data = [
    ('Random Forest',
     0.5093, 0.0073, 0.5100, 0.77, 0.43, 0.33,
     '+0.0745', 1,
     'BEST model. max_depth=None, min_samples_leaf=4, n_estimators=200 (4/5 folds)'),
    ('XGBoost',
     0.4820, 0.0039, 0.4799, 0.73, 0.39, 0.31,
     '+0.0444', 2,
     'n_estimators=200, max_depth=6, learning_rate=0.1 (all 5 folds agreed)'),
    ('Linear SVM',
     0.4680, 0.0034, 0.4693, 0.77, 0.34, 0.29,
     '+0.0338', 3,
     'C=0.1; strong Good/Bad recall balance. class_weight=balanced'),
    ('LR + SMOTE',
     0.4619, 0.0064, 0.4648, 0.73, 0.39, 0.27,
     '+0.0293', 4,
     'Oversampling inside CV pipeline (ImbPipeline). C=0.1'),
    ('KNN',
     0.4462, 0.0066, 0.4516, 0.77, 0.38, 0.21,
     '+0.0161', 5,
     'Manhattan, k=11, distance weights. 15K stratified subsample. No class_weight.'),
    ('LR + class_weight (Baseline)',
     0.4385, 0.0034, 0.4355, 0.70, 0.33, 0.28,
     '---', 6,
     'BASELINE. C=0.1, L2, lbfgs. class_weight=balanced'),
    ('LR + ROS',
     0.4376, 0.0038, 0.4352, 0.70, 0.33, 0.28,
     '-0.0003', 7,
     'ROS gives no improvement over class_weight. C=0.01'),
]

ROW_COLORS1 = [
    'C6EFCE',  # green   – best
    'DDEBF7',  # blue
    'DDEBF7',
    'DDEBF7',
    'DDEBF7',
    'FFEB9C',  # yellow  – baseline
    'FCE4D6',  # red     – below baseline
]

for idx, row_data in enumerate(models_data):
    row_num = idx + 3
    model, cv_mean, cv_std, test_f1, good_f1, mixed_f1, bad_f1, vs_lr, rank, notes = row_data
    values = [model, cv_mean, cv_std, test_f1, good_f1, mixed_f1, bad_f1, vs_lr, rank, notes]
    write_row(ws1, row_num, values, fill_hex=ROW_COLORS1[idx])
    ws1.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws1.cell(row=row_num, column=10).alignment = Alignment(
        horizontal='left', vertical='center', wrap_text=True)
    for col in [2, 3, 4, 5, 6, 7]:
        ws1.cell(row=row_num, column=col).number_format = '0.0000'

ws1.conditional_formatting.add(
    'D3:D9',
    ColorScaleRule(
        start_type='min',  start_color='FCE4D6',
        mid_type='percentile', mid_value=50, mid_color='FFEB9C',
        end_type='max',    end_color='C6EFCE'
    )
)

# Spacer + findings
for r in [10, 11]:
    ws1.row_dimensions[r].height = 10
add_finding_box(ws1, 12,
    'Key Takeaways:  Random Forest (0.5100) is the best model, beating XGBoost (+0.0301), '
    'Linear SVM (+0.0407), and the LR baseline (+0.0745). '
    'Tree-based models (RF, XGBoost) outperform linear models because they capture non-linear feature interactions. '
    'KNN beats the baseline (+0.016) but lags behind ensemble methods and has the lowest Bad F1 (0.21) '
    'due to the lack of class_weight support. '
    'LR + ROS (-0.0003) is essentially identical to the baseline, confirming that duplicating minority samples adds no information.',
    NCOLS1, height=65)

set_col_widths(ws1, [26, 14, 10, 12, 12, 12, 12, 14, 7, 46])


# =============================================================================
# SHEET 2 — Imbalance Handling (LR variants)
# =============================================================================
ws2 = wb.create_sheet('Imbalance Handling')
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 22
ws2.row_dimensions[3].height = 42

NCOLS2 = 10
add_title(ws2, 1,
    'Logistic Regression  |  Imbalance Handling Strategy Comparison  (class_weight vs ROS vs SMOTE)',
    NCOLS2, '375623')

ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS2)
sub2 = ws2.cell(row=2, column=1,
    value='All three variants use Logistic Regression (L2, lbfgs) on the same 0.5-threshold dataset. '
          'Only the imbalance handling mechanism differs.')
sub2.fill      = hfill('E2EFDA')
sub2.font      = Font(italic=True, size=10, color='375623')
sub2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
sub2.border    = thin_border()

HEADERS2 = [
    'Strategy',
    'CV Macro F1\n(mean)',
    'CV Std\n(+/-)',
    'Test\nMacro F1',
    'Good F1\n(test)',
    'Mixed F1\n(test)',
    'Bad F1\n(test)',
    'vs Baseline\n(+/-)',
    'Best C\n(final model)',
    'Mechanism',
]
write_header(ws2, 3, HEADERS2, '375623')

imbalance_data = [
    ('LR + SMOTE',
     0.4619, 0.0064, 0.4648, 0.73, 0.39, 0.27,
     '+0.0293', 'C=0.1',
     'Synthetic minority oversampling (interpolation). Applied inside ImbPipeline '
     'to prevent leakage into validation fold.'),
    ('LR + class_weight=balanced  (BASELINE)',
     0.4385, 0.0034, 0.4355, 0.70, 0.33, 0.28,
     '---', 'C=0.1',
     'Loss function re-weighting proportional to inverse class frequency. '
     'No data augmentation, no leakage risk.'),
    ('LR + ROS',
     0.4376, 0.0038, 0.4352, 0.70, 0.33, 0.28,
     '-0.0003', 'C=0.01',
     'Random duplication of minority class samples. Applied inside ImbPipeline. '
     'No new information added.'),
]

ROW_COLORS2 = ['C6EFCE', 'FFEB9C', 'FCE4D6']

for idx, row_data in enumerate(imbalance_data):
    row_num = idx + 4
    strat, cv_mean, cv_std, test_f1, good_f1, mixed_f1, bad_f1, vs_base, best_c, mechanism = row_data
    values = [strat, cv_mean, cv_std, test_f1, good_f1, mixed_f1, bad_f1, vs_base, best_c, mechanism]
    write_row(ws2, row_num, values, fill_hex=ROW_COLORS2[idx])
    ws2.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws2.cell(row=row_num, column=10).alignment = Alignment(
        horizontal='left', vertical='center', wrap_text=True)
    for col in [2, 3, 4, 5, 6, 7]:
        ws2.cell(row=row_num, column=col).number_format = '0.0000'
    ws2.row_dimensions[row_num].height = 35

ws2.conditional_formatting.add(
    'D4:D6',
    ColorScaleRule(
        start_type='min', start_color='FCE4D6',
        end_type='max',   end_color='C6EFCE'
    )
)

for r in [7, 8]:
    ws2.row_dimensions[r].height = 10
add_finding_box(ws2, 9,
    'Key Findings:  SMOTE is the only strategy that meaningfully improves over the baseline (+0.029 Macro F1). '
    'It creates synthetic minority samples via interpolation, giving the model genuinely new patterns to learn. '
    'ROS gives no meaningful gain (-0.0003) — duplicating existing samples adds no new information. '
    'class_weight=balanced is simpler, faster, and essentially equivalent to ROS; it is the preferred default. '
    'SMOTE improves Mixed F1 (0.33->0.39) and Good F1 (0.70->0.73), but Bad F1 is slightly lower (0.27 vs 0.28) '
    'possibly due to interpolated Bad samples introducing noise near the Mixed boundary.',
    NCOLS2, height=70)

set_col_widths(ws2, [30, 14, 10, 12, 12, 12, 12, 16, 14, 46])


# =============================================================================
# SHEET 3 — Threshold Experiment (0.5 vs 0.4)
# =============================================================================
ws3 = wb.create_sheet('Threshold Experiment')
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[2].height = 22
ws3.row_dimensions[3].height = 42

NCOLS3 = 11
add_title(ws3, 1,
    'Threshold Experiment  |  Original 0.5-Boundary vs Steam Natural 0.4-Boundary  |  Model: Logistic Regression',
    NCOLS3, '7030A0')

ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS3)
sub3 = ws3.cell(row=2, column=1,
    value='Same 147 features, same 80/20 stratified split, same LR configuration. '
          'Only the class boundary thresholds change.')
sub3.fill      = hfill('EAD1DC')
sub3.font      = Font(italic=True, size=10, color='7030A0')
sub3.alignment = Alignment(horizontal='center', vertical='center')
sub3.border    = thin_border()

HEADERS3 = [
    'Threshold',
    'Good\nBoundary',
    'Mixed\nBoundary',
    'Bad\nBoundary',
    'Good %\n(train)',
    'Mixed %\n(train)',
    'Bad %\n(train)',
    'CV Macro F1',
    'CV Std\n(+/-)',
    'Test\nMacro F1',
    'vs 0.5\nThreshold',
]
write_header(ws3, 3, HEADERS3, '7030A0')

threshold_data = [
    ('0.5 — Original\n(project standard)',
     '>= 75%', '50 – 74%', '< 50%',
     '63.3%', '28.3%', '8.4%',
     0.4385, 0.0034, 0.4355, '---'),
    ('0.4 — Steam Natural\n(this experiment)',
     '>= 70%', '40 – 69%', '< 40%',
     '71.4%', '24.3%', '4.3%',
     0.4173, 0.0027, 0.4135, '-0.0220'),
]

ROW_COLORS3 = ['C6EFCE', 'FCE4D6']

for idx, row_data in enumerate(threshold_data):
    row_num = idx + 4
    write_row(ws3, row_num, list(row_data), fill_hex=ROW_COLORS3[idx])
    ws3.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws3.cell(row=row_num, column=8).number_format  = '0.0000'
    ws3.cell(row=row_num, column=9).number_format  = '0.0000'
    ws3.cell(row=row_num, column=10).number_format = '0.0000'
    ws3.row_dimensions[row_num].height = 38

# Delta row
ws3.row_dimensions[6].height = 22
write_row(ws3, 6,
    ['Delta (0.4 - 0.5)', '', '', '', '+8.1 pp', '-4.0 pp', '-4.1 pp',
     -0.0212, -0.0007, -0.0220, ''],
    fill_hex='F2F2F2')
ws3.cell(row=6, column=1).font = Font(bold=True, size=11, color='C00000')
for col in [8, 9, 10]:
    cell = ws3.cell(row=6, column=col)
    cell.font         = Font(bold=True, size=11, color='C00000')
    cell.number_format = '0.0000'

# Per-class F1 breakdown
ws3.row_dimensions[7].height = 12
ws3.row_dimensions[8].height = 28
write_header(ws3, 8,
    ['', 'Good F1\n(test)', 'Mixed F1\n(test)', 'Bad F1\n(test)',
     'Good Support\n(test)', 'Mixed Support\n(test)', 'Bad Support\n(test)',
     '', '', '', ''],
    '7030A0')

per_class_data = [
    ('0.5 — Original',  0.70, 0.33, 0.28, '7,168 (63.3%)', '3,208 (28.3%)', '955 (8.4%)',   '', '', '', ''),
    ('0.4 — Experiment', 0.72, 0.35, 0.17, '8,086 (71.4%)', '2,754 (24.3%)', '491 (4.3%)',   '', '', '', ''),
    ('Change',           '+0.02', '+0.02', '-0.11', '', '', '', '', '', '', ''),
]
ROW_COLORS3b = ['C6EFCE', 'FCE4D6', 'F2F2F2']

for idx, row_data in enumerate(per_class_data):
    row_num = idx + 9
    write_row(ws3, row_num, list(row_data), fill_hex=ROW_COLORS3b[idx])
    ws3.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    if idx == 2:
        for col in [2, 3, 4]:
            ws3.cell(row=row_num, column=col).font = Font(bold=True, size=11, color='C00000')

for r in [12, 13]:
    ws3.row_dimensions[r].height = 10
add_finding_box(ws3, 14,
    'Key Findings:  The 0.4-threshold (Steam natural boundaries) performs WORSE overall (-0.022 Macro F1). '
    'The dominant cause is Bad class collapse: Bad shrinks from 8.4% to 4.3% of the dataset, '
    'leaving only 1,963 training samples at <40%. Bad F1 drops from 0.28 to 0.17 (precision 0.10 on test set). '
    'Good and Mixed improve marginally (+0.02 each) but cannot offset the Bad collapse. '
    'Conclusion: the original 0.5-threshold provides a more balanced Bad class with sufficient training signal. '
    'The 0.4-boundary aligns with Steam display labels but is inferior for ML classification.',
    NCOLS3, height=70)

set_col_widths(ws3, [22, 12, 14, 12, 13, 13, 12, 14, 10, 14, 16])


# =============================================================================
out_path = r'c:\Users\I764176\Documents\University\Semester 3\Data Mining\Model_Results_Comparison.xlsx'
wb.save(out_path)
print('Saved:', out_path)
